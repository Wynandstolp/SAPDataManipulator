import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def print_time(description):
    print(description + datetime.now().strftime("%H:%M:%S"))


read_path = "Tests/MujaList_2.xlsx"
write_path = "Tests/MujaList_2_material_linked_v2.xlsx"

print_time("finished imports: ")
df = pd.read_excel(read_path, sheet_name="CT BOM")
print_time("finished read_excel: ")
wb = load_workbook(filename=read_path, data_only=True)
print_time("finished load_workbook: ")
ws = wb.worksheets[0]
print_time("finished set worksheet: ")

list_dict = {}

for index, row in df.iterrows():
    if row['Construction type'] in list_dict.keys():
        list_dict[row['Construction type']].extend([row['Material']])
        # remove duplicates
        arr = list(set(list_dict[row['Construction type']]))
        list_dict[row['Construction type']] = arr
    else:
        list_dict[row['Construction type']] = [row['Material']]

print_time("finished dataframe: ")

for row in range(3, 56944):  # 56944
    result_cell = 'AP{}'.format(row)
    print('Row: ', row)
    try:
        constructType = list_dict[ws['E{}'.format(row)].value]
        if constructType != None:
            column_min = 42
            column_max = len(constructType) + 42
            for column in range(column_min, column_max):
                ws.cell(
                    row=row, column=column).value = constructType[column - column_min]
        else:
            ws[result_cell] = ""
            break
    except Exception as ex:
        ws[result_cell] = "No Mat# linked"
        continue

wb.save(write_path)
